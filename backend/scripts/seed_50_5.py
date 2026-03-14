"""
Seed MongoDB with 5 providers and 50 patients (10 patients per provider).
Generates an Excel file with usernames, passwords, and provider-patient mapping.

Run from backend folder:
  pip install openpyxl   # if not installed
  python -m scripts.seed_50_5

Password for all seeded users: HealthGuard1
"""
import os
import random
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path

from dotenv import load_dotenv
from pymongo import MongoClient
import bcrypt

# Load .env from backend folder
load_dotenv(Path(__file__).resolve().parent.parent / ".env")

MONGODB_URL = os.getenv("MONGODB_URL", "mongodb://localhost:27017")
DATABASE_NAME = os.getenv("DATABASE_NAME", "healthguard")
SEED_PASSWORD = "HealthGuard1"
NUM_PROVIDERS = 5
NUM_PATIENTS = 50
PATIENTS_PER_PROVIDER = 10  # 5 * 10 = 50

# Reuse sample data from seed_data
from scripts.seed_data import (
    BLOOD_TYPES,
    FIRST_NAMES,
    GENDERS,
    GOAL_TARGETS,
    GOAL_TYPES,
    GOAL_UNITS,
    HOSPITALS,
    LAST_NAMES,
    REMINDER_CATEGORIES,
    REMINDER_STATUSES,
    REMINDER_TITLES,
    SAMPLE_ALLERGIES,
    SAMPLE_MEDS,
    SPECIALIZATIONS,
)


def utc_now():
    return datetime.now(timezone.utc)


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _date_to_datetime(d: date) -> datetime:
    return datetime.combine(d, time.min, tzinfo=timezone.utc)


def seed_wellness_goals_and_reminders(db, patient_user_ids, now):
    today = date.today()
    for patient_id in patient_user_ids:
        for goal_type in random.sample(GOAL_TYPES, random.randint(1, 3)):
            low, high = GOAL_TARGETS[goal_type]
            target = round(random.uniform(low, high), 1) if isinstance(low, float) else float(random.randint(low, high))
            goal_doc = {
                "patient_id": patient_id,
                "goal_type": goal_type,
                "target_value": target,
                "unit": GOAL_UNITS[goal_type],
                "is_active": True,
                "created_at": now,
            }
            gr = db["wellness_goals"].insert_one(goal_doc)
            goal_id = gr.inserted_id
            for d in range(random.randint(7, 14)):
                log_date = today - timedelta(days=d)
                logged = random.randint(int(target * 0.5), int(target * 1.1)) if goal_type == "steps" else round(random.uniform(target - 1, target + 1), 1)
                db["goal_logs"].insert_one({
                    "goal_id": goal_id,
                    "patient_id": patient_id,
                    "logged_value": logged,
                    "log_date": _date_to_datetime(log_date),
                    "created_at": now,
                })
        for _ in range(random.randint(2, 4)):
            title = random.choice(REMINDER_TITLES)
            due_date = today + timedelta(days=random.randint(-20, 45))
            db["reminders"].insert_one({
                "patient_id": patient_id,
                "title": title,
                "description": f"Schedule or complete: {title}",
                "due_date": _date_to_datetime(due_date),
                "status": random.choice(REMINDER_STATUSES),
                "category": random.choice(REMINDER_CATEGORIES),
                "created_at": now,
            })


def main():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl")
        return

    client = MongoClient(MONGODB_URL, serverSelectionTimeoutMS=60_000, connectTimeoutMS=60_000)
    db = client[DATABASE_NAME]

    now = utc_now()
    password_hash = hash_password(SEED_PASSWORD)

    # Build provider list with emails that won't conflict with main seed
    provider_emails = [f"demo_provider_{i + 1}@healthguard.demo" for i in range(NUM_PROVIDERS)]
    patient_emails = [f"demo_patient_{i + 1}@healthguard.demo" for i in range(NUM_PATIENTS)]

    # Check for duplicate emails
    existing = set()
    for doc in db["users"].find({"email": {"$in": provider_emails + patient_emails}}, {"email": 1}):
        existing.add(doc["email"])
    if existing:
        print(f"These emails already exist in DB. Delete them first or skip: {existing}")
        return

    provider_records = []  # (user_id, email, full_name, specialization)
    provider_user_ids = []

    for i in range(NUM_PROVIDERS):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        email = provider_emails[i]
        license_num = f"MD{20000 + i}{random.randint(10, 99)}"
        specialization = SPECIALIZATIONS[i % len(SPECIALIZATIONS)]
        user_doc = {
            "email": email,
            "password_hash": password_hash,
            "role": "provider",
            "is_active": True,
            "consent_given": True,
            "created_at": now,
            "updated_at": now,
        }
        r = db["users"].insert_one(user_doc)
        provider_user_ids.append(r.inserted_id)
        full_name = f"{first} {last}"
        provider_records.append((r.inserted_id, email, full_name, specialization))
        db["provider_profiles"].insert_one({
            "user_id": r.inserted_id,
            "full_name": full_name,
            "phone": f"+1-555-{400 + i}{random.randint(10, 99)}-{random.randint(1000, 9999)}",
            "specialization": specialization,
            "license_number": license_num,
            "hospital_or_clinic": HOSPITALS[i % len(HOSPITALS)],
            "years_of_experience": random.randint(5, 25),
            "updated_at": now,
        })

    patient_records = []  # (user_id, email, full_name, provider_email)
    patient_user_ids = []

    for i in range(NUM_PATIENTS):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        email = patient_emails[i]
        provider_idx = i // PATIENTS_PER_PROVIDER
        provider_email = provider_emails[provider_idx]
        user_doc = {
            "email": email,
            "password_hash": password_hash,
            "role": "patient",
            "is_active": True,
            "consent_given": True,
            "created_at": now,
            "updated_at": now,
        }
        r = db["users"].insert_one(user_doc)
        patient_user_ids.append(r.inserted_id)
        full_name = f"{first} {last}"
        patient_records.append((r.inserted_id, email, full_name, provider_email))
        allergies = [a for a in random.sample(SAMPLE_ALLERGIES, random.randint(0, 2)) if a != "None"]
        meds = [m for m in random.sample(SAMPLE_MEDS, random.randint(0, 2)) if m != "None"]
        db["patient_profiles"].insert_one({
            "user_id": r.inserted_id,
            "full_name": full_name,
            "age": random.randint(22, 75),
            "gender": random.choice(GENDERS),
            "phone": f"+1-555-{500 + (i % 50)}{random.randint(10, 99)}-{random.randint(1000, 9999)}",
            "allergies": allergies,
            "current_medications": meds,
            "blood_type": random.choice(BLOOD_TYPES),
            "emergency_contact": f"+1-555-600-{random.randint(1000, 9999)}",
            "updated_at": now,
        })
        db["provider_patient"].insert_one({
            "provider_id": provider_user_ids[provider_idx],
            "patient_id": r.inserted_id,
            "assigned_at": now,
            "is_active": True,
        })

    seed_wellness_goals_and_reminders(db, patient_user_ids, now)

    # Build provider email -> full name for Excel
    provider_email_to_name = {email: name for (_, email, name, _) in provider_records}

    # Write Excel
    wb = Workbook()
    ws_prov = wb.active
    ws_prov.title = "Providers"
    ws_prov.append(["Username (Email)", "Password", "Full Name", "Specialization", "Patients Assigned"])
    for (_, email, full_name, specialization) in provider_records:
        ws_prov.append([email, SEED_PASSWORD, full_name, specialization, PATIENTS_PER_PROVIDER])

    ws_pat = wb.create_sheet("Patients", 1)
    ws_pat.append(["Username (Email)", "Password", "Full Name", "Assigned Provider (Email)", "Assigned Provider Name"])
    for (_, email, full_name, prov_email) in patient_records:
        ws_pat.append([email, SEED_PASSWORD, full_name, prov_email, provider_email_to_name[prov_email]])

    ws_mapping = wb.create_sheet("Mapping Summary", 2)
    ws_mapping.append(["Provider Email", "Provider Name", "Patient Count", "Patient Emails"])
    for (_, prov_email, prov_name, _) in provider_records:
        patients_for_provider = [pr[1] for pr in patient_records if pr[3] == prov_email]
        ws_mapping.append([prov_email, prov_name, len(patients_for_provider), ", ".join(patients_for_provider)])

    out_path = Path(__file__).resolve().parent.parent / "healthguard_user_mapping_50_5.xlsx"
    wb.save(out_path)

    print(f"Seeded {NUM_PROVIDERS} providers and {NUM_PATIENTS} patients (10 per provider).")
    print(f"Password for all: {SEED_PASSWORD}")
    print(f"Excel saved: {out_path}")


if __name__ == "__main__":
    main()
